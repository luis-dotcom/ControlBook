import pandas as pd
import numpy as np
import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Alignment
from Dados import *

class ControlBook():

    def __init__(self):
        self.planilha_report = pd.read_excel("file.xlsx")

    def create_controlbook(self):

        self.planilha_report['Ticket Externo'] = self.planilha_report['Ticket Externo'].fillna('0')
        self.planilha_report['Ticket Externo'] =  self.planilha_report['Ticket Externo'].astype(str)
        #AJUSTES PARA TIPO DATE-TIME
        self.planilha_report['Data de Normalização'] = pd.to_datetime(self.planilha_report['Data de Normalização'])
        self.planilha_report['Aberto em'] = pd.to_datetime(self.planilha_report['Aberto em'])
        self.planilha_report['Resolvido em'] = pd.to_datetime(self.planilha_report['Resolvido em'])
        self.planilha_report['Fechado em'] = pd.to_datetime(self.planilha_report['Fechado em'])

        #FORMATAÇÃO DATE-TIME
        self.planilha_report['Aberto em'] = self.planilha_report['Aberto em'].dt.strftime("%d/%m/%Y %H:%M")
        self.planilha_report['Data de Normalização'] = self.planilha_report['Data de Normalização'].dt.strftime("%d/%m/%Y %H:%M")
        self.planilha_report['Resolvido em'] = self.planilha_report['Resolvido em'].dt.strftime("%d/%m/%Y %H:%M")
        self.planilha_report['Fechado em'] = self.planilha_report['Fechado em'].dt.strftime("%d/%m/%Y %H:%M")
        self.planilha_report['Horário do Incidente'] = self.planilha_report['Horário do Incidente'].dt.strftime("%d/%m/%Y %H:%M")

        self.planilha_report['Indicador do SLA de Solução'] = self.planilha_report['Indicador do SLA de Solução'].fillna('--')

        # 1º IF = VERIFICA VALORES VÁZIOS E APAGA TODAS AS COLUNAS CASO ESTEJAM VAZIAS E CRIA UMA NOVA
        if self.planilha_report.loc[:, "Causa Raiz de ToIP"].any() == False and self.planilha_report.loc[:,
                                                                        "Causa Raiz de Dados"].any() == False and self.planilha_report.loc[
                                                                                                                    :,
                                                                                                                    "Causa Raiz de Vídeo"].any() == False:
            self.planilha_report = self.planilha_report.drop(columns=['Causa Raiz de ToIP', 'Causa Raiz de Dados', 'Causa Raiz de Vídeo'])
            self.planilha_report.insert(16, 'Causa Raiz', '')
            
            
        # 2º ELSE = VERIFICA SE EXISTEM VALORES NAS COLUNAS DE CAUSA RAIZ TOIP, VÍDEO E DADOS, SE SIM COLOCAR EM UMA ÚNICA COLUNA CHAMADA CAUSA RAIZ  
        else:
            self.planilha_report[["Causa Raiz de ToIP", "Causa Raiz de Dados", "Causa Raiz de Vídeo"]] = self.planilha_report[
                ["Causa Raiz de ToIP", "Causa Raiz de Dados", "Causa Raiz de Vídeo"]].fillna('')
            dados = self.planilha_report[["Causa Raiz de ToIP", "Causa Raiz de Dados", "Causa Raiz de Vídeo"]].values

            dados_causa = []

            for i in range(len(dados)):
                dados_causa.append(''.join(dados[i]))
                i += 1
            causa = pd.DataFrame(dados_causa, columns=['Causa Raiz'])
            self.planilha_report = pd.concat([self.planilha_report, causa], axis=1)
            self.planilha_report = self.planilha_report.drop(columns=['Causa Raiz de ToIP', 'Causa Raiz de Dados', 'Causa Raiz de Vídeo'])


            
        #CRIAR COLUNA UNIDADE NA POSIÇÃO 13
        if "Unidade" not in self.planilha_report.columns:
            self.planilha_report.insert(13, 'Unidade', '')


        #ORGANIZAÇÃO DAS COLUNAS NA ORDEM DESCRITA
        self.planilha_report = self.planilha_report[Dados().colunas]

        #VERIFICA SE O CHAMADO É INCIDENTE OU REQUISIÇÃO E FAZ A FORMATAÇÃO
        self.planilha_report.loc[self.planilha_report['Ticket Externo'] == '0', 'Ticket Externo'] = '--'
        self.planilha_report[['Resolvido em', 'Fechado em', 'Resolução']] = self.planilha_report[
            ['Resolvido em', 'Fechado em', 'Resolução']].fillna('EM ABERTO')
        self.planilha_report[['Horário do Incidente', 'Data de Normalização']] = self.planilha_report.loc[
            self.planilha_report['Categoria'] != 'Solicitação de serviço', ['Horário do Incidente', 'Data de Normalização']].fillna(
            'EM ABERTO')
        self.planilha_report.loc[self.planilha_report['Categoria'] != 'Incidente', 'Horário do Incidente'] = '--'
        self.planilha_report.loc[self.planilha_report['Categoria'] != 'Incidente', 'Data de Normalização'] = '--'
        self.planilha_report.loc[self.planilha_report['Categoria'] != 'Incidente', 'Causa Raiz'] = '--'

        incidentes = self.planilha_report.loc[self.planilha_report['Categoria'] != 'Solicitação de serviço', 'Causa Raiz'].index.tolist()
        planilha = self.planilha_report['Causa Raiz'].values

        for i in incidentes:
            if planilha[i] == '':
                self.planilha_report.loc[i, 'Causa Raiz'] = 'EM ABERTO'

                
        residentes = self.planilha_report.loc[self.planilha_report['Criado por'] == 'Residentes BNB', 'Responsável: Equipe'].index.tolist()
        residentes_dados = self.planilha_report['Responsável: Equipe'].values

        for i in residentes:
            if residentes_dados[i] == 'Residentes BNB':
                self.planilha_report.loc[i, 'Responsável: Equipe'] = 'BNB RESIDENTE'
                self.planilha_report.loc[i, 'Cliente (Organização)'] = 'BNB'

        self.planilha_report.loc[self.planilha_report['Cliente (Organização)'] == 'Veneza', 'Cliente (Organização)' ] = 'GRUPO VENEZA'
                
        #FORMATA TEXTO DO TIPO DE SERVIÇO PARA O PADRÃO CONTROLBOOK

        if self.planilha_report['Serviço (Completo)'].loc[self.planilha_report['Serviço (Completo)'] == 'ToIP » Solicitações » Topologia/Relatório/Análise de ToIP'].any() == True:
                self.planilha_report['Serviço (Completo)'].loc[self.planilha_report['Serviço (Completo)'] == 'ToIP » Solicitações » Topologia/Relatório/Análise de ToIP'] ='TOPOLOGIA/RELATÓRIO/ANÁLISE DE TOIP'

        if self.planilha_report['Serviço (Completo)'].loc[self.planilha_report['Serviço (Completo)'] == 'ToIP » Incidentes » Inoperância Total de ToIP'].any() == True:
                self.planilha_report['Serviço (Completo)'].loc[self.planilha_report['Serviço (Completo)'] == 'ToIP » Incidentes » Inoperância Total de ToIP'] ='INOPERÂNCIA TOTAL DE TOIP'


        dados = Dados().dados
        for i in dados:
            self.planilha_report['Serviço (Completo)'] = self.planilha_report['Serviço (Completo)'].str.replace(i, '')

        

        #FORMATAÇÃO DE HORA 
        self.planilha_report["Tempo de vida (Descontando parado - Horas corridas)"] = self.planilha_report[
            "Tempo de vida (Descontando parado - Horas corridas)"].apply(lambda t: dt.time(t.hour, t.minute, t.second))
        self.planilha_report["Tempo de vida (Horas corridas)"] = self.planilha_report["Tempo de vida (Horas corridas)"].apply(
            lambda t: dt.time(t.hour, t.minute, t.second))


        
        #CONVERTE TODAS AS COLUNAS PARA TEXTO
        self.planilha_report[Dados().colunas] = self.planilha_report[Dados().colunas].astype(str)

        #COLOCA TODAS AS LETRAS MAIUSCULAS
        for col in Dados().colunas:
            self.planilha_report[col] = self.planilha_report[col].str.upper()

        self.planilha_report['Indicador do SLA de Solução'] = self.planilha_report['Indicador do SLA de Solução'].replace(np.nan,'--')

        self.planilha_report['Unidade'].loc[self.planilha_report['Cliente (Organização)'] == 'SEBRAE'] = 'CEARÁ'


        for i in Dados().dados_gtw_sebrae:
            self.planilha_report['Unidade'].loc[self.planilha_report['Assunto'] == Dados().gtw_sebrae[i]] = i


        #RECONHCER UNIDADE RECIFE PARA BNB E BASETELCO
        self.planilha_report['Unidade'].loc[self.planilha_report['Cliente (Organização)'] == 'BASE TELCO'] = 'RECIFE'
        self.planilha_report['Unidade'].loc[self.planilha_report['Cliente (Organização)'] == 'BNB'] = 'FORTALEZA'


        unidades_clientes = self.planilha_report['Assunto']
        unidades_clientes = unidades_clientes.values
        c = []
        for i in unidades_clientes:
            c.append(i)
        unidades_clientes = []
        for i in c:
            unidades_clientes.append(i)

        for i in Dados().clientes_unidades:
            for v in range(len(unidades_clientes)):
                if i in unidades_clientes[v]:
                    self.planilha_report['Unidade'][unidades_clientes.index(unidades_clientes[v])] = i


        #REGRA PARA APAGAR CHAMADOS CANCELADOS E CLIENTES NÃO LISTADOS    
       
        self.planilha_report['Categoria'] = self.planilha_report['Categoria'].fillna('')
        indexname = self.planilha_report.loc[self.planilha_report['Categoria'] == ''].index
        self.planilha_report.drop(indexname, inplace=True)

        self.planilha_report['Cliente (Organização)'] = self.planilha_report['Cliente (Organização)'].fillna('PREFEITURA DE LAGOA SANTA')
        indexnamee = self.planilha_report.loc[self.planilha_report['Cliente (Organização)'] == 'PREFEITURA DE LAGOA SANTA'].index
        self.planilha_report.drop(indexnamee, inplace=True)

        self.planilha_report['Cliente (Organização)'] = self.planilha_report['Cliente (Organização)'].fillna('PREFEITURA DE NOVA LIMA')
        indexnamee = self.planilha_report.loc[self.planilha_report['Cliente (Organização)'] == 'PREFEITURA DE NOVA LIMA'].index
        self.planilha_report.drop(indexnamee, inplace=True)

        self.planilha_report['Cliente (Organização)'] = self.planilha_report['Cliente (Organização)'].fillna('BASE MOBILE')
        indexnamee = self.planilha_report.loc[self.planilha_report['Cliente (Organização)'] == 'BASE MOBILE'].index
        self.planilha_report.drop(indexnamee, inplace=True)

        self.planilha_report['Cliente (Organização)'] = self.planilha_report['Cliente (Organização)'].fillna('IAUPE')
        indexnamee = self.planilha_report.loc[self.planilha_report['Cliente (Organização)'] == 'IAUPE'].index
        self.planilha_report.drop(indexnamee, inplace=True)

        self.planilha_report['Responsável: Equipe'] = self.planilha_report['Responsável: Equipe'].fillna('CONECTA +')
        indexnamee = self.planilha_report.loc[self.planilha_report['Responsável: Equipe'] == 'CONECTA +'].index
        self.planilha_report.drop(indexnamee, inplace=True)
        
        self.planilha_report = self.planilha_report.sort_values(by='Número')


        self.planilha_report = self.planilha_report.rename(columns={'Número':'NÚMERO','Categoria':'CATEGORIA','Origem de Abertura':'ORIGEM DE ABERTURA','Serviço (Completo)': 'SERVIÇO', 'Aberto em':'ABERTO EM',
                                                           'Horário do Incidente': 'HORÁRIO DO INCIDENTE','Data de Normalização':'DATA DE NORMALIZAÇÃO','Resolvido em':'RESOLVIDO EM','Fechado em':'FECHADO EM ',
                                                           'Criado por':'CRIADO POR','Serviço (1º Nível)':'SERVIÇO (1º NÍVEL)','Assunto':'ASSUNTO','Descrição':'DESCRIÇÃO','Cliente (Organização)':'CLIENTE (ORGANIZAÇÃO)',
                                                           'Unidade':'UNIDADE','Responsável: Equipe':'RESPONSÁVEL: EQUIPE','Causa Raiz':'CAUSA RAIZ','Resolução':'RESOLUÇÃO','Indicador do SLA de Solução':'INDICADOR DO SLA DE SOLUÇÃO',
                                                           'Tempo parado (Horas corridas)':'TEMPO PARADO (HORAS CORRIDAS)','Tempo de vida (Descontando parado - Horas corridas)':'TEMPO DE VIDA (DESCONTANDO PARADO - HORAS CORRIDAS)',
                                                           'Tempo de vida (Horas corridas)':'TEMPO DE VIDA (HORAS CORRIDAS)','Ticket Externo':'TICKET EXTERNO'})
        #GERAR O ARQUIVO CONTROLBOOK
    def paint(self):    
        self.planilha_report.to_excel('ControlBook.xlsx', index=False)

        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        wb = load_workbook(filename='ControlBook.xlsx')
        ws = wb.active
        e = ws['A1':'A'+ str(ws.max_row)]
        a = ws.min_column
        for b in range(ws.max_column):
            for i in range(len(e)):
                ws.cell(row=i + 1, column=b+1).border = thin_border
        for row in ws[1:ws.max_row]:
            for position in row:
                cell = position
                cell.alignment = Alignment(horizontal='center')
        yellow = PatternFill(fill_type='solid', start_color='00FFFF00', end_color='00FFFF00') #COR DAS CELULAS
        linha = 2
        for i in range(ws.max_row):
            data = [ws.cell(row=linha, column=i).value for i in range(1,24)] # PEGANDO TODOS AS LINHAS E COLUNAS ENTRE AS COLUNAS A E W COLENTANDO OS DADOS POR LINHAS
        
            if "EM ABERTO" in data:
                for row in ws[f'A{linha}:W{linha}']:
                    for cell in row:
                        cell.fill = yellow
            linha+=1
        wb.save('ControlBook.xlsx')

gerar = ControlBook()

gerar.create_controlbook()
gerar.paint()